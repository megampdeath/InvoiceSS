from __future__ import annotations

import csv
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.db.models import Invoice
from app.invoices.service import party_to_dict, supplier_party

EXPORT_COLUMNS = [
    "invoice_id",
    "status",
    "supplier_name",
    "supplier_vat_number",
    "invoice_number",
    "invoice_date",
    "due_date",
    "currency",
    "subtotal_amount",
    "tax_amount",
    "total_amount",
    "iban",
    "payment_terms",
    "original_filename",
    "uploaded_at",
    "reviewed_at",
    "warnings",
]


def build_export(db: Session, invoices: list[Invoice], export_format: str) -> tuple[str, bytes, int]:
    timestamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    rows = [invoice_export_row(invoice) for invoice in invoices]
    if export_format == "csv":
        return f"invoices-export-{timestamp}.csv", _build_csv(rows), len(rows)
    return f"invoices-export-{timestamp}.xlsx", _build_xlsx(rows), len(rows)


def invoice_export_row(invoice: Invoice) -> dict[str, object]:
    supplier = party_to_dict(supplier_party(invoice))
    return {
        "invoice_id": invoice.id,
        "status": invoice.status,
        "supplier_name": supplier.get("name"),
        "supplier_vat_number": supplier.get("vat_number"),
        "invoice_number": invoice.invoice_number,
        "invoice_date": invoice.invoice_date.isoformat() if invoice.invoice_date else "",
        "due_date": invoice.due_date.isoformat() if invoice.due_date else "",
        "currency": invoice.currency,
        "subtotal_amount": money(invoice.subtotal_amount),
        "tax_amount": money(invoice.tax_amount),
        "total_amount": money(invoice.total_amount),
        "iban": invoice.iban,
        "payment_terms": invoice.payment_terms,
        "original_filename": invoice.original_filename,
        "uploaded_at": invoice.created_at.isoformat(),
        "reviewed_at": invoice.reviewed_at.isoformat() if invoice.reviewed_at else "",
        "warnings": ";".join(warning.code for warning in invoice.warnings),
    }


def _build_csv(rows: list[dict[str, object]]) -> bytes:
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8")


def _build_xlsx(rows: list[dict[str, object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoices"
    sheet.append(EXPORT_COLUMNS)
    header_fill = PatternFill(fill_type="solid", fgColor="E7EEF8")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
    for row in rows:
        sheet.append([row.get(column) for column in EXPORT_COLUMNS])
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)
    money_columns = ["I", "J", "K"]
    for column in money_columns:
        for cell in sheet[column][1:]:
            cell.number_format = "#,##0.00"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def money(value: Decimal | None) -> str:
    return f"{value:.2f}" if value is not None else ""
